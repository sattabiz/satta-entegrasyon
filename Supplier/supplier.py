import shutil
from openpyxl import load_workbook
from Common.qt_compat import Qt
from Common.qt_compat import (
    QAbstractItemView,
    QCheckBox,
    QFileDialog,
    QHeaderView,
    QHBoxLayout,
    QLabel,
    QLineEdit,
    QMessageBox,
    QPushButton,
    QTableWidget,
    QTableWidgetItem,
    QVBoxLayout,
    QWidget,
)
from Supplier.supplier_reader import SupplierReader, SupplierReaderConfig
import json
from Supplier.push_suppliers import SattaSupplierPushConnector
from Common.path_helper import project_path
import unicodedata


class SupplierSendTab(QWidget):
    def __init__(self):
        super().__init__()

        root_layout = QVBoxLayout(self)

        title_label = QLabel("Tedarikçi Gönderim Ekranı")

        title_row = QHBoxLayout()
        title_row.addWidget(title_label)
        title_row.addStretch()

        self.load_button = QPushButton("Tedarikçileri Al")
        self.template_button = QPushButton("Tedarikçi Şablonunu İndir")
        self.import_template_button = QPushButton("Şablondan Tedarikçi Yükle")
        self.send_button = QPushButton("Seçili Tedarikçileri Satta'ya Gönder")
        self.edit_table_checkbox = QCheckBox("Tabloyu düzenlenebilir yap")
        self.edit_table_checkbox.toggled.connect(self.toggle_table_edit_mode)
        self.load_button.clicked.connect(self.load_suppliers)
        self.template_button.clicked.connect(self.download_supplier_template)
        self.import_template_button.clicked.connect(self.import_suppliers_from_template)
        self.send_button.clicked.connect(self.send_selected_suppliers)
        title_row.addWidget(self.load_button)
        title_row.addWidget(self.template_button)
        title_row.addWidget(self.import_template_button)
        title_row.addWidget(self.send_button)
        title_row.addWidget(self.edit_table_checkbox)
        root_layout.addLayout(title_row)

        self.search_input = QLineEdit()
        self.search_input.setMinimumHeight(36)
        self.search_input.setMinimumWidth(320)
        self.search_input.setPlaceholderText("Tedarikçi kodu veya adı")

        self.search_button = QPushButton("🔍")
        self.search_button.setMinimumHeight(36)
        self.search_button.setMinimumWidth(44)

        search_row = QHBoxLayout()
        search_row.addWidget(self.search_input)
        search_row.addWidget(self.search_button)

        root_layout.addLayout(search_row)

        self.supplier_table = QTableWidget(0, 7)
        self.supplier_table.setHorizontalHeaderLabels([
            "Seç",
            "Kod",
            "Tedarikçi Adı",
            "İlgili Kişi",
            "Telefon Numarası",
            "E-posta Adresi",
            "Vergi No",
        ])
        self.supplier_table.setSelectionBehavior(QAbstractItemView.SelectRows)

        supplier_header = self.supplier_table.horizontalHeader()
        supplier_header.setSectionResizeMode(QHeaderView.Interactive)

        self.supplier_table.setColumnWidth(0, 44)
        self.supplier_table.setColumnWidth(1, 110)
        self.supplier_table.setColumnWidth(2, 220)
        self.supplier_table.setColumnWidth(3, 160)
        self.supplier_table.setColumnWidth(4, 140)
        self.supplier_table.setColumnWidth(5, 220)
        self.supplier_table.setColumnWidth(6, 130)
        self.supplier_table.setEditTriggers(QAbstractItemView.NoEditTriggers)
        root_layout.addWidget(self.supplier_table)

        status_info_layout = QHBoxLayout()

        self.selected_info_label = QLabel("Seçili tedarikçi sayısı: 0")

        status_info_layout.addWidget(self.selected_info_label)

        root_layout.addLayout(status_info_layout)

        self.all_suppliers = []
        self.search_button.clicked.connect(self.run_search_with_feedback)
        self.search_input.returnPressed.connect(self.run_search_with_feedback)
        self.search_input.textChanged.connect(self.filter_suppliers)

    def download_supplier_template(self):
        template_path = project_path("Templates", "supplierTemplate.xlsx")

        if not template_path.exists():
            QMessageBox.critical(self, "Şablon Bulunamadı", "Tedarikçi Excel şablonu bulunamadı.")
            return

        save_path, _ = QFileDialog.getSaveFileName(
            self,
            "Tedarikçi Şablonunu Kaydet",
            "supplierTemplate.xlsx",
            "Excel Dosyaları (*.xlsx)",
        )

        if not save_path:
            return

        if not save_path.lower().endswith(".xlsx"):
            save_path += ".xlsx"

        try:
            shutil.copyfile(template_path, save_path)
        except OSError as exc:
            QMessageBox.critical(self, "Kopyalama Hatası", f"Şablon dosyası kaydedilemedi:\n{exc}")
            return

        QMessageBox.information(self, "Şablon Hazır", "Tedarikçi şablonu başarıyla kaydedildi.")

    def run_search_with_feedback(self):
        self.filter_suppliers(show_no_results_message=True)

    def fetch_suppliers(self):
        try:
            from Common.path_helper import user_data_path
            app_settings_file = user_data_path("app_settings.json")
            if app_settings_file.exists():
                with open(app_settings_file, "r", encoding="utf-8") as f:
                    settings = json.load(f)
                
                logo_settings = settings.get("logo", {})
                config = SupplierReaderConfig(
                    server=logo_settings.get("server", ""),
                    database=logo_settings.get("database", ""),
                    username=logo_settings.get("username", ""),
                    password=logo_settings.get("password", ""),
                    firm_no=int(logo_settings.get("firm_no", "")),
                    period_no=int(logo_settings.get("period_no", ""))
                )
            else:
                config = SupplierReaderConfig()

            reader = SupplierReader(config)
            return reader.get_suppliers()
            
        except Exception as e:
            QMessageBox.critical(self, "Veritabanı Hatası", f"Tedarikçileri çekerken hata oluştu:\n{str(e)}")
            return []

    def apply_supplier_data(self, rows):
        self.all_suppliers = rows
        self.populate_supplier_table(self.all_suppliers)

        try:
            self.supplier_table.itemChanged.disconnect(self.update_selected_count)
        except RuntimeError:
            pass
        except TypeError:
            pass

        self.supplier_table.itemChanged.connect(self.update_selected_count)
        self.update_status_summary()

        if self.supplier_table.rowCount() > 0:
            self.supplier_table.selectRow(0)

    def load_suppliers(self):
        supplier_rows = self.fetch_suppliers()
        self.apply_supplier_data(supplier_rows)

    def populate_supplier_table(self, rows):
        self.supplier_table.setRowCount(0)
        for row_data in rows:
            row_index = self.supplier_table.rowCount()
            self.supplier_table.insertRow(row_index)

            select_item = QTableWidgetItem()
            select_item.setFlags(Qt.ItemIsUserCheckable | Qt.ItemIsEnabled | Qt.ItemIsSelectable)
            select_item.setCheckState(Qt.Unchecked)
            select_item.setText("")
            self.supplier_table.setItem(row_index, 0, select_item)

            for col_index, value in enumerate(row_data, start=1):
                item = QTableWidgetItem(value)
                if not self.edit_table_checkbox.isChecked() or col_index == 1:
                    item.setFlags(item.flags() & ~Qt.ItemIsEditable)
                self.supplier_table.setItem(row_index, col_index, item)

    def filter_suppliers(self, *_args, show_no_results_message=False):
        search_text = self.search_input.text().strip().lower()

        if not search_text:
            self.populate_supplier_table(self.all_suppliers)
            self.update_status_summary()
            if self.supplier_table.rowCount() > 0:
                self.supplier_table.selectRow(0)
            return

        filtered_rows = []
        for row in self.all_suppliers:
            supplier_code = str(row[0]).lower()
            supplier_name = str(row[1]).lower()

            if search_text in supplier_code or search_text in supplier_name:
                filtered_rows.append(row)

        self.populate_supplier_table(filtered_rows)
        self.update_status_summary()

        if self.supplier_table.rowCount() > 0:
            self.supplier_table.selectRow(0)
        elif search_text and show_no_results_message:
            QMessageBox.information(self, "Arama Sonucu", "Aramaya uygun tedarikçi bulunamadı.")

    def get_selected_suppliers(self):
        selected_suppliers = []
        invalid_suppliers = []

        for row in range(self.supplier_table.rowCount()):
            check_item = self.supplier_table.item(row, 0)
            if check_item is None or check_item.checkState() != Qt.Checked:
                continue

            code_item = self.supplier_table.item(row, 1)
            name_item = self.supplier_table.item(row, 2)
            person_item = self.supplier_table.item(row, 3)
            phone_item = self.supplier_table.item(row, 4)
            email_item = self.supplier_table.item(row, 5)
            tax_id_item = self.supplier_table.item(row, 6)

            supplier_code = code_item.text().strip() if code_item else ""
            supplier_name = name_item.text().strip() if name_item else ""
            invited_person = person_item.text().strip() if person_item else ""
            phone = phone_item.text().strip() if phone_item else ""
            invited_email = email_item.text().strip() if email_item else ""
            tax_id = tax_id_item.text().strip() if tax_id_item else ""

            missing_fields = []
            if not supplier_name:
                missing_fields.append("Tedarikçi Adı")
            if not invited_person:
                missing_fields.append("İlgili Kişi")
            if not invited_email:
                missing_fields.append("E-posta Adresi")
            if not tax_id:
                missing_fields.append("Vergi No")
            if not supplier_code:
                missing_fields.append("Kod")
            if not phone:
                missing_fields.append("Telefon")

            if missing_fields:
                supplier_label = supplier_name or supplier_code or f"Satır {row + 1}"
                invalid_suppliers.append(f"{supplier_label} -> Eksik alanlar: {', '.join(missing_fields)}")
                continue

            selected_suppliers.append(
                {
                    "supplier_name": supplier_name,
                    "invited_person": invited_person,
                    "phone": phone,
                    "invited_email": invited_email,
                    "tax_id": tax_id,
                    "erp_id": supplier_code,
                }
            )

        return selected_suppliers, invalid_suppliers

    def send_selected_suppliers(self):
        selected_suppliers, invalid_suppliers = self.get_selected_suppliers()

        if invalid_suppliers:
            missing_text = "\n".join(invalid_suppliers)
            QMessageBox.warning(
                self,
                "Eksik Zorunlu Alan",
                f"Aşağıdaki tedarikçiler gönderilmedi çünkü zorunlu alanlar boş:\n{missing_text}",
            )

        if not selected_suppliers:
            QMessageBox.warning(self, "Seçim Yok", "Gönderilecek geçerli tedarikçi bulunamadı.")
            return

        connector = SattaSupplierPushConnector()

        try:
            connector.push_suppliers(selected_suppliers)
        except Exception as exc:
            QMessageBox.critical(self, "Aktarım Hatası", f"Seçili tedarikçiler Satta'ya gönderilemedi:\n{exc}")
            return

        QMessageBox.information(
            self,
            "Aktarım Tamamlandı",
            f"Seçili {len(selected_suppliers)} tedarikçi Satta'ya gönderildi.",
        )

    def update_selected_count(self):
        selected_count = 0
        for row in range(self.supplier_table.rowCount()):
            item = self.supplier_table.item(row, 0)
            if item is not None and item.checkState() == Qt.Checked:
                selected_count += 1

        self.selected_info_label.setText(f"Seçili tedarikçi sayısı: {selected_count}")

    def update_status_summary(self):
        self.update_selected_count()

    def toggle_table_edit_mode(self, checked):
        if checked:
            self.supplier_table.setEditTriggers(
                QAbstractItemView.DoubleClicked
                | QAbstractItemView.EditKeyPressed
                | QAbstractItemView.SelectedClicked
            )
        else:
            self.supplier_table.setEditTriggers(QAbstractItemView.NoEditTriggers)

        current_rows = []
        for row in range(self.supplier_table.rowCount()):
            row_data = []
            for col in range(1, self.supplier_table.columnCount()):
                item = self.supplier_table.item(row, col)
                row_data.append(item.text() if item else "")
            current_rows.append(tuple(row_data))

        self.populate_supplier_table(current_rows)

    def normalize_header(self, value):
        normalized_text = unicodedata.normalize("NFKD", str(value or "").strip().casefold())
        normalized_text = "".join(
            character for character in normalized_text if not unicodedata.combining(character)
        )
        normalized_text = normalized_text.replace("_", " ").replace("-", " ")
        return " ".join(normalized_text.split())

    def import_suppliers_from_template(self):
        file_path, _ = QFileDialog.getOpenFileName(
            self,
            "Tedarikçi Şablonunu Seç",
            "",
            "Excel Dosyaları (*.xlsx)",
        )

        if not file_path:
            return

        try:
            workbook = load_workbook(file_path, data_only=True)
            worksheet = workbook.active
        except Exception as exc:
            QMessageBox.critical(self, "Şablon Okuma Hatası", f"Excel şablonu okunamadı:\n{exc}")
            return

        rows = list(worksheet.iter_rows(values_only=True))
        if not rows:
            QMessageBox.warning(self, "Boş Dosya", "Seçilen Excel dosyasında veri bulunamadı.")
            return

        header_row = rows[0]
        normalized_headers = [self.normalize_header(cell) for cell in header_row]

        header_aliases = {
            "code": {"kod", "erp_id", "erp id", "erpid", "supplier code", "tedarikçi kodu", "tedarikci kodu"},
            "supplier_name": {"tedarikçi adı", "tedarikci adı", "tedarikci adi", "tedarikçi adi", "name", "supplier name", "unvan", "firma", "firma adı"},
            "contact": {"ilgili kişi", "ilgili kisi", "contact", "contact person", "yetkili"},
            "phone": {"telefon", "telefon numarası", "telefon numarasi", "phone", "gsm"},
            "email": {"e-posta", "e posta", "email", "eposta", "mail"},
            "tax_id": {"vergi no", "vergi numarası", "vergi numarasi", "tax id", "tax number", "vkn"},
        }

        column_map = {}
        for index, header in enumerate(normalized_headers):
            for field_name, aliases in header_aliases.items():
                if header in aliases and field_name not in column_map:
                    column_map[field_name] = index

        has_header_mapping = len(column_map) >= 4
        data_rows = rows[1:] if has_header_mapping else rows

        imported_rows = []
        for row in data_rows:
            if row is None:
                continue

            row_values = list(row)
            if not any(str(cell).strip() for cell in row_values if cell is not None):
                continue

            if has_header_mapping:
                def cell_value(field_name):
                    cell_index = column_map.get(field_name)
                    if cell_index is None or cell_index >= len(row_values):
                        return ""
                    return str(row_values[cell_index] or "").strip()

                supplier_row = (
                    cell_value("code"),
                    cell_value("supplier_name"),
                    cell_value("contact"),
                    cell_value("phone"),
                    cell_value("email"),
                    cell_value("tax_id"),
                )
            else:
                padded_values = [str(cell or "").strip() for cell in row_values[:6]]
                while len(padded_values) < 6:
                    padded_values.append("")
                supplier_row = tuple(padded_values)

            if any(value.strip() for value in supplier_row):
                imported_rows.append(supplier_row)

        if not imported_rows:
            QMessageBox.warning(self, "Veri Bulunamadı", "Excel dosyasında içe aktarılacak uygun tedarikçi verisi bulunamadı.")
            return

        self.apply_supplier_data(imported_rows)
        QMessageBox.information(self, "İçe Aktarım Tamamlandı", f"{len(imported_rows)} tedarikçi şablondan yüklendi.")